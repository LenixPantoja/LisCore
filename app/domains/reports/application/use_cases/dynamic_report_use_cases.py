import asyncio
import base64
import io
import re
import tempfile
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import requests as _requests
from fastapi import HTTPException, status
from jinja2 import Template, TemplateError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.domains.reports.domain.dynamic_report_models import DynamicReport, ReportParameter
from app.domains.reports.infrastructure.dynamic_report_repository import DynamicReportRepository

# ── Logo ─────────────────────────────────────────────────────────────────────
_LOGO_PATH = (
    Path(__file__).parent.parent.parent / "infrastructure" / "templates" / "marca_agua.png"
)


def _get_logo_base64() -> str:
    """Return the company logo as an inline base64 data URI, or empty string on failure."""
    try:
        with open(_LOGO_PATH, "rb") as f:
            encoded = base64.b64encode(f.read()).decode("utf-8")
        return f"data:image/png;base64,{encoded}"
    except Exception:
        return ""


# ── PDF header ────────────────────────────────────────────────────────────────

def _inject_header_into_html(html: str, header_html: str) -> str:
    """Inject the PDF header right after <body> tag in the existing HTML."""
    match = re.search(r"(<body[^>]*>)", html, re.IGNORECASE)
    if match:
        insert_pos = match.end()
        return html[:insert_pos] + "\n" + header_html + "\n" + html[insert_pos:]
    # Fallback: prepend header
    return header_html + html


def _build_pdf_header(
    report_name: str,
    dr_code: Optional[str],
    dr_version: Optional[str],
    dr_emission_date: Optional[str],
) -> str:
    """Build the standard 3-column PDF header using table layout."""
    logo_src = _get_logo_base64()
    logo_tag = (
        f'<img src="{logo_src}" style="height:38px; vertical-align:middle;" />'
        if logo_src
        else ""
    )
    code_val = dr_code or "—"
    version_val = dr_version or "—"
    emission_val = dr_emission_date or "—"

    return f"""
<table style="width:100%; border-collapse:collapse; margin-bottom:0;" cellspacing="0" cellpadding="6">
  <tr>
    <td style="width:22%; vertical-align:middle; padding:6px 8px;">
      {logo_tag}
      <div style="color:#233248; font-weight:bold; font-size:12pt; margin-top:2px;">Liscore</div>
      <div style="color:#4a5568; font-size:7pt; font-style:italic;">Sistema de gestion de laboratorio</div>
    </td>
    <td style="text-align:center; vertical-align:middle; padding:6px 8px;">
      <span style="font-weight:bold; font-size:10.5pt; letter-spacing:0.4pt; text-transform:uppercase;">
        {report_name}
      </span>
    </td>
    <td style="width:20%; vertical-align:middle; padding:0;">
      <table style="border:1px solid #233248; font-size:8pt; width:100%; border-collapse:collapse;" cellspacing="0" cellpadding="4">
        <tr><td style="border-bottom:1px solid #233248; padding:3px 6px;">
          <span style="color:#4a5568;">C&oacute;digo: </span><b>{code_val}</b>
        </td></tr>
        <tr><td style="border-bottom:1px solid #233248; padding:3px 6px;">
          <span style="color:#4a5568;">Versi&oacute;n: </span><b>{version_val}</b>
        </td></tr>
        <tr><td style="padding:3px 6px;">
          <span style="color:#4a5568;">Emisi&oacute;n: </span><b>{emission_val}</b>
        </td></tr>
      </table>
    </td>
  </tr>
</table>
<table style="width:100%; border-collapse:collapse; margin-bottom:12px;" cellspacing="0" cellpadding="0">
  <tr><td style="height:3px; background:#233248;"></td></tr>
</table>
"""


def _build_complete_pdf_html(
    header_html: str,
    body_content: str,
    page_size: str = "carta",
    orientation: str = "portrait",
) -> str:
    """Return a complete, self-contained HTML document ready for Gotenberg."""
    _size_map = {"carta": "letter", "oficio": "legal"}
    css_size = _size_map.get(page_size, "letter")

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8" />
  <style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: Helvetica, Arial, sans-serif; font-size: 9pt; color: #1e293b; }}
    @page {{ size: {css_size} {orientation}; margin: 1.5cm 1.5cm 2cm 1.5cm; }}
    table {{ border-collapse: collapse; width: 100%; }}
  </style>
</head>
<body>
{header_html}
{body_content}
</body>
</html>"""


# ── Gotenberg ─────────────────────────────────────────────────────────────────

def _call_gotenberg_sync(html: str, page_size: str, orientation: str) -> bytes:
    """Write HTML to a temp file, POST it to Gotenberg, delete the file, return PDF bytes."""
    # Paper dimensions in inches
    # Carta: 8.5 x 11  |  Oficio (legal): 8.5 x 14
    size_map = {
        "carta": (8.5, 11.0),
        "oficio": (8.5, 14.0),
        "legal": (8.5, 14.0),
    }
    paper_w, paper_h = size_map.get(page_size.lower(), (8.5, 11.0))
    if orientation == "landscape":
        paper_w, paper_h = paper_h, paper_w

    # Write HTML to a temporary file
    tmp_dir = Path(tempfile.gettempdir())
    tmp_file = tmp_dir / f"report_{uuid.uuid4().hex}.html"
    tmp_file.write_text(html, encoding="utf-8")

    url = f"{settings.GOTENBERG_URL.rstrip('/')}/forms/chromium/convert/html"
    data = {
        "paperWidth": str(paper_w),
        "paperHeight": str(paper_h),
        "marginTop": "0.5",
        "marginBottom": "0.5",
        "marginLeft": "0.5",
        "marginRight": "0.5",
        "scale": "1.0",
        "printBackground": "true",
    }
    try:
        with tmp_file.open("rb") as fh:
            files = {"file": ("index.html", fh, "text/html")}
            response = _requests.post(url, files=files, data=data, timeout=30)
        response.raise_for_status()
        return response.content
    except _requests.exceptions.ConnectionError as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"No se pudo conectar al servicio Gotenberg en {settings.GOTENBERG_URL}. "
                   f"Verifique que el contenedor esté corriendo.",
        ) from exc
    except _requests.exceptions.HTTPError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Gotenberg devolvió un error: {exc.response.status_code} — {exc.response.text[:200]}",
        ) from exc
    except _requests.exceptions.Timeout as exc:
        raise HTTPException(
            status_code=status.HTTP_504_GATEWAY_TIMEOUT,
            detail="Gotenberg tardó demasiado en responder (timeout 30s).",
        ) from exc
    finally:
        # Always remove the temporary HTML file
        try:
            tmp_file.unlink(missing_ok=True)
        except OSError:
            pass


async def _html_to_pdf_bytes(html: str, page_size: str, orientation: str) -> bytes:
    """Async wrapper: runs the Gotenberg HTTP call in a thread pool to avoid blocking."""
    return await asyncio.get_event_loop().run_in_executor(
        None, _call_gotenberg_sync, html, page_size, orientation
    )


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

async def _resolve_parameter_options(
    db: AsyncSession, param: ReportParameter
) -> Optional[List[Dict[str, Any]]]:
    """If the parameter has a source_query, execute it and return option list."""
    if param.rp_source_query:
        rows = await DynamicReportRepository.execute_source_query(db, param.rp_source_query)
        # Expect columns: value, label (first two columns as fallback)
        options = []
        for row in rows:
            keys = list(row.keys())
            options.append({
                "value": row[keys[0]],
                "label": row[keys[1]] if len(keys) > 1 else str(row[keys[0]]),
            })
        return options
    return None


def _render_html(template_str: str, data: List[dict], params: dict) -> str:
    """Render a Jinja2 HTML template with report data and filter params."""
    try:
        template = Template(template_str)
        return template.render(data=data, params=params)
    except TemplateError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error al renderizar la plantilla HTML: {exc}",
        )


def _coerce_params(
    raw_params: dict, param_definitions: List[ReportParameter]
) -> dict:
    """Convert raw string values to proper Python types based on rp_type definitions.

    asyncpg requires typed Python objects (date, datetime, int, float) — not plain strings.
    """
    type_map = {p.rp_name: p.rp_type for p in param_definitions}
    coerced: dict = {}

    for key, value in raw_params.items():
        if value is None:
            coerced[key] = value
            continue

        param_type = type_map.get(key, "text")

        if param_type == "date" and isinstance(value, str):
            try:
                coerced[key] = date.fromisoformat(value)
            except ValueError:
                coerced[key] = value

        elif param_type == "datetime" and isinstance(value, str):
            try:
                coerced[key] = datetime.fromisoformat(value)
            except ValueError:
                coerced[key] = value

        elif param_type == "number" and isinstance(value, str):
            try:
                coerced[key] = int(value) if "." not in value else float(value)
            except ValueError:
                coerced[key] = value

        elif param_type in ("select", "multiselect") and isinstance(value, str):
            # If the value looks like an integer (e.g. a foreign-key id), coerce it
            # so asyncpg does not receive a string where the DB column is INTEGER.
            try:
                coerced[key] = int(value)
            except ValueError:
                try:
                    coerced[key] = float(value)
                except ValueError:
                    coerced[key] = value  # leave as text (e.g. a code string)

        else:
            coerced[key] = value

    return coerced


# ---------------------------------------------------------------------------
# Use cases
# ---------------------------------------------------------------------------

async def list_reports(db: AsyncSession) -> List[DynamicReport]:
    return list(await DynamicReportRepository.list_all(db))


async def list_reports_tree(db: AsyncSession) -> List[Dict[str, Any]]:
    """Return active reports grouped by category as a tree structure.

    Reports without a category appear under the key ``None``.
    Within each category, reports are sorted alphabetically.
    """
    reports = await DynamicReportRepository.list_all(db)

    # Preserve insertion order so categories appear in the order they are first seen
    groups: Dict[Optional[str], List[Dict[str, Any]]] = {}
    for report in reports:
        category = report.dr_category_name or None
        if category not in groups:
            groups[category] = []
        groups[category].append({
            "dr_id": report.dr_id,
            "dr_name": report.dr_name,
            "dr_description": report.dr_description,
            "dr_active": report.dr_active,
        })

    return [{"category": cat, "reports": items} for cat, items in groups.items()]


async def get_report_detail(db: AsyncSession, report_id: int) -> Dict[str, Any]:
    """Return the report metadata together with each parameter options resolved."""
    report = await DynamicReportRepository.get_by_id(db, report_id)

    resolved_params = []
    for param in report.parameters:
        entry: Dict[str, Any] = {
            "rp_id": param.rp_id,
            "rp_name": param.rp_name,
            "rp_label": param.rp_label,
            "rp_type": param.rp_type,
            "rp_required": param.rp_required,
            "rp_default_value": param.rp_default_value,
            "rp_order_index": param.rp_order_index,
            "options": None,
        }
        if param.rp_type in ("select", "multiselect"):
            entry["options"] = await _resolve_parameter_options(db, param)
        resolved_params.append(entry)

    return {
        "dr_id": report.dr_id,
        "dr_name": report.dr_name,
        "dr_description": report.dr_description,
        "dr_active": report.dr_active,
        "parameters": resolved_params,
    }


async def create_report(
    db: AsyncSession,
    report_data: dict,
    params_data: List[dict],
) -> DynamicReport:
    report = await DynamicReportRepository.create(db, report_data, params_data)
    await db.commit()
    await db.refresh(report)
    return report


async def update_report(
    db: AsyncSession,
    report_id: int,
    report_data: dict,
    params_data: Optional[List[dict]],
) -> DynamicReport:
    report = await DynamicReportRepository.update(db, report_id, report_data, params_data)
    await db.commit()
    await db.refresh(report)
    return report


async def delete_report(db: AsyncSession, report_id: int) -> None:
    await DynamicReportRepository.soft_delete(db, report_id)
    await db.commit()


async def run_report(
    db: AsyncSession,
    report_id: int,
    filter_params: dict,
) -> Dict[str, Any]:
    """Execute the report SQL with the provided filters and render HTML."""
    report = await DynamicReportRepository.get_by_id(db, report_id)

    # Validate required parameters are present
    for param in report.parameters:
        if param.rp_required and param.rp_name not in filter_params:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"El par\u00e1metro requerido '{param.rp_label}' no fue enviado.",
            )

    # Coerce string values to proper Python types for asyncpg
    typed_params = _coerce_params(filter_params, report.parameters)

    data_rows = await DynamicReportRepository.execute_report_query(
        db, report.dr_sql_query, typed_params
    )
    html = _render_html(report.dr_html_template, data_rows, filter_params)

    return {
        "report_id": report.dr_id,
        "report_name": report.dr_name,
        "total_rows": len(data_rows),
        "html": html,
    }


async def export_report_pdf(
    db: AsyncSession,
    report_id: int,
    filter_params: dict,
    page_size: str = "carta",
    orientation: str = "portrait",
) -> Dict[str, Any]:
    """Run the report and convert the rendered HTML to PDF via Gotenberg (base64-encoded).

    The PDF includes a standard 3-column header:
      [Logo + brand] | [Report title centered] | [C\u00f3digo / Versi\u00f3n / Emisi\u00f3n box]

    Args:
        page_size: 'carta' (letter) or 'oficio' (legal).
        orientation: 'portrait' or 'landscape'.
    """
    result = await run_report(db, report_id, filter_params)

    # Use exactly the same HTML that /run returns — preserves all design and styles
    pdf_bytes = await _html_to_pdf_bytes(result["html"], page_size, orientation)

    encoded = base64.b64encode(pdf_bytes).decode("utf-8")
    filename = f"reporte_{report_id}_{result['report_name'].replace(' ', '_')}.pdf"

    return {
        "filename": filename,
        "base64_pdf": encoded,
        "report_name": result["report_name"],
        "total_rows": result["total_rows"],
    }


async def export_report_xlsx(
    db: AsyncSession,
    report_id: int,
    filter_params: dict,
) -> Dict[str, Any]:
    """Run the report SQL and export the raw data rows as an XLSX file (base64-encoded)."""
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="openpyxl no est\u00e1 instalado. No es posible exportar a XLSX.",
        ) from exc

    report = await DynamicReportRepository.get_by_id(db, report_id)

    # Validate required parameters
    for param in report.parameters:
        if param.rp_required and param.rp_name not in filter_params:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"El par\u00e1metro requerido '{param.rp_label}' no fue enviado.",
            )

    typed_params = _coerce_params(filter_params, report.parameters)
    data_rows = await DynamicReportRepository.execute_report_query(
        db, report.dr_sql_query, typed_params
    )

    # Build XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report.dr_name[:31]  # Sheet name max 31 chars

    if data_rows:
        headers = list(data_rows[0].keys())

        # Header row styling
        header_fill = PatternFill(start_color="233248", end_color="233248", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Data rows
        for row_idx, row in enumerate(data_rows, start=2):
            for col_idx, key in enumerate(headers, start=1):
                value = row[key]
                # Convert date/datetime to string for Excel compatibility
                if hasattr(value, "isoformat"):
                    value = value.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        for col in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded = base64.b64encode(buf.getvalue()).decode("utf-8")
    filename = f"reporte_{report_id}_{report.dr_name.replace(' ', '_')}.xlsx"

    return {
        "filename": filename,
        "base64_xlsx": encoded,
        "report_name": report.dr_name,
        "total_rows": len(data_rows),
    }